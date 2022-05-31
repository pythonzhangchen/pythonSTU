# -*- coding: utf-8 -*-
# @Time : 2022/4/28 15:31 
# @Author : chen.zhang 
# @File : prac_excel.py
import openpyxl as ex

# from openpyxl import Workbook   # 创建新的excel
# from openpyxl import load_workbook  # 加载现存的

# # 对表的创建操作
# wb = ex.Workbook()  # 实例化 wb workbook
#
# ws = wb.active  # 选择当激活的sheet
#
# ws.title = 'my first'  # Sheet excel, Sheet1 Sheet2
#
# ws = wb.create_sheet('ccc')  # 可以单独创建一个Sheet
#
# ws = wb['my first']  # 选择需要操作的sheet名
#
# ws['A1'] = 100  # 基于单元格位置写入数据
# ws['A2'] = 200  # 这个就跟我们去做字典赋值没有区别
#
# ws.cell(row=1, column=1, value=100000)  # 基于单元格坐标写入数据
#
# wb.save('excels/myExcel1.xlsx')

# # 对表的基本写入操作 第一种
# courses = [
#     ['Python', '趣讲精炼'],
#     ['Python', 'Django'],
#     ['职场', 'PPT高手'],
#     ['K12', '语数外'],
# ]
#
# wb = ex.Workbook()  # 实例化 workbook
# ws = wb.active  # 激活其中一个sheet work sheet
# ws.title = 'myFirstSheet'   # 创建一个sheet名字
# ws['A1'] = '类型'
# ws['B1'] = '课程名字'
#
# for data in courses:
#     # print(data)
#     ws.append(data)
# wb.save('excels/myExcel2.xlsx')


# 对表的基本写入操作 第二种 采用加载现有的excel，然后再次写入
new_course = [
    ['Python2', '趣讲精炼2'],
    ['Python2', 'Django2'],
    ['职场2', 'PPT高手2'],
    ['K12 2', '语数外2'],
]
wb = ex.load_workbook('excels/myExcel2.xlsx')
ws = wb['myFirstSheet']

# for x in ws.rows:  # 对行进行输出
#     print(x[0].value, x[1].value)

print('*' * 20)  # 黄金分割线

# for col in ws.columns:  # 对列进行输出
#     for v in col:
#         print(v.value)

print('当前sheet共计行数：', ws.max_row)  # 注意如果你要添加  从第五行开始添加

for new_data in new_course:
    course_type, course_name = new_data  # 双赋值
    print(course_type, course_name)
    row_counter = ws.max_row + 1  # 特别注意这里的问题。 ws max 每次都会检测  必须统一一下
    ws.cell(row=row_counter, column=1, value=course_type)
    ws.cell(row=row_counter, column=2, value=course_name)

wb.save('excels/myExcel2.xlsx')
